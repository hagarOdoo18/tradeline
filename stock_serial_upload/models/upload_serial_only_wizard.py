# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError
from collections import OrderedDict
import base64
import openpyxl
from io import BytesIO


class UploadSerialOnlyWizard(models.TransientModel):
    _name = 'upload.serial.only.wizard'
    _description = 'Upload Serials for Internal Transfers (Serial Column Only)'

    file = fields.Binary(string='Excel File', required=True)
    filename = fields.Char(string='File Name')
    picking_id = fields.Many2one('stock.picking', string='Transfer', required=True)

    def action_upload_serials(self):
        if not self.file:
            raise UserError(_('Please upload an Excel file.'))

        picking = self.picking_id
        if picking.picking_type_code != 'internal':
            raise UserError(_('This wizard only works for Internal Transfers.'))
        if picking.state != 'draft':
            raise UserError(_('This wizard only works for Draft Transfer.'))

        try:
            data = base64.b64decode(self.file)
            wb = openpyxl.load_workbook(filename=BytesIO(data), data_only=True)
            sheet = wb.active
        except Exception as e:
            raise UserError(_('Invalid Excel file. Must be .xlsx format.\n\nError: %s') % str(e))

        headers = [str(cell.value).strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_map = {name: idx for idx, name in enumerate(headers) if name}

        if 'serial' not in header_map:
            raise UserError(_('Excel must contain a column named "serial".'))

        # ------------------------------------------------------------------
        # 1) Read the file first and group VALID serials by product, so each
        #    product gets ONE move with the correct demand (instead of
        #    bumping product_uom_qty one row at a time).
        # ------------------------------------------------------------------
        lots_by_product = OrderedDict()   # product -> [lot, lot, ...]
        seen_serials = set()
        errors = []

        for row in sheet.iter_rows(min_row=2):
            serial_name = str(row[header_map['serial']].value or '').strip()
            if not serial_name:
                continue
            if serial_name in seen_serials:
                # same serial listed twice in the file -> ignore the duplicate
                continue
            seen_serials.add(serial_name)

            lot = self.env['stock.lot'].search([
                ('name', '=', serial_name),
                ('location_id', '=', picking.location_id.id),
            ], limit=1)

            if not lot:
                errors.append(_('Not Found serial %s at this stock') % serial_name)
                continue

            lots_by_product.setdefault(lot.product_id, []).append(lot)

        if errors:
            raise UserError(_('Upload aborted. Fix these and try again:\n\n') + '\n'.join(errors))

        # ------------------------------------------------------------------
        # 2) One move per product, one move line per serial.
        #    Odoo 18: stock.move.line has NO qty_done / reserved_uom_qty.
        #      - 'quantity' = the reserved quantity for this line
        #      - 'picked'   = mark the line as done
        #    Filling 'quantity' up to the move demand means the move is fully
        #    reserved, so "Check Availability" / confirm will NOT generate its
        #    own extra reservation lines.
        # ------------------------------------------------------------------
        MoveLine = self.env['stock.move.line']
        processed = 0

        for product, lots in lots_by_product.items():
            move = picking.move_ids_without_package.filtered(
                lambda mv: mv.product_id == product)[:1]

            if not move:
                move = self.env['stock.move'].create({
                    'picking_id': picking.id,
                    'product_id': product.id,
                    'name': product.display_name,
                    'product_uom_qty': len(lots),
                    'product_uom': product.uom_id.id,
                    'location_id': picking.location_id.id,
                    'location_dest_id': picking.location_dest_id.id,
                    'company_id': picking.company_id.id,
                })
            else:
                move.product_uom_qty += len(lots)

            for lot in lots:
                # safe re-run: skip a serial already on this move
                if move.move_line_ids.filtered(lambda ml: ml.lot_id == lot):
                    continue

                MoveLine.create({
                    'picking_id': picking.id,
                    'move_id': move.id,
                    'product_id': product.id,
                    'product_uom_id': product.uom_id.id,
                    'lot_id': lot.id,
                    'quantity': 1.0,   # Odoo 18: reserved quantity for this line
                    'picked': True,    # Odoo 18: mark the line as done
                    'location_id': move.location_id.id,
                    'location_dest_id': move.location_dest_id.id,
                })
                processed += 1

        return {'type': 'ir.actions.client', 'tag': 'reload'}
