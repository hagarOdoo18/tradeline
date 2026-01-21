# -*- coding: utf-8 -*-
from odoo import models, fields
from odoo.exceptions import UserError
import base64
import openpyxl
from io import BytesIO
import logging

_logger = logging.getLogger(__name__)


class UploadBarcodeWizard(models.TransientModel):
    _name = 'upload.product.barcode.wizard'
    _description = 'Upload Product Variant Barcodes from Excel'

    file = fields.Binary(string='Excel File', required=True)
    filename = fields.Char(string='File Name')

    def action_upload_barcodes(self):
        if not self.file:
            raise UserError('Please upload an Excel file.')

        try:
            data = base64.b64decode(self.file)
            wb = openpyxl.load_workbook(filename=BytesIO(data), data_only=True)
            sheet = wb.active
        except Exception as e:
            raise UserError(f'Invalid Excel file.\nError: {e}')

        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_map = {str(h).strip().lower(): i for i, h in enumerate(headers) if h}

        required = ['barcode', 'e_code', 'gs1_code']
        for col in required:
            if col not in header_map:
                raise UserError(f"Missing column '{col}' in Excel file.")

        updated = 0
        not_found = []
        errors = []

        Product = self.env['product.product']

        for row in sheet.iter_rows(min_row=2):
            with self.env.cr.savepoint():  # isolate each row
                try:

                    barcode = str(row[header_map['barcode']].value or '').strip()
                    e_code = str(row[header_map['e_code']].value or '').strip()
                    gs1_code = str(row[header_map['gs1_code']].value or '').strip()


                    domain = [('barcode', '=', barcode)]

                    product = Product.search(domain, limit=1)

                    if not product:
                        not_found.append(f"{barcode} ")
                        continue

                    product.write({
                        'barcode': barcode or '',
                        'e_invoicing_code': e_code or '',
                        'gs1_code': gs1_code or '',
                    })
                    updated += 1

                except Exception as e:
                     # rollback current row only
                    errors.append(f"Error updating {barcode}: {str(e)}")
                    _logger.warning(f"Failed to update {barcode}: {e}")
                    continue

        # Build summary
        summary = [f"✅ Upload completed.\nUpdated variants: {updated}"]

        if not_found:
            summary.append("\n🟠 Products not found:")
            summary.extend(not_found)
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Barcode Upload Complete',
                    'message': '\n'.join(summary[:20]) + ('\n...(truncated)' if len(summary) > 20 else ''),
                    'type': 'success' if not errors else 'warning',
                    'sticky': False,
                },
            }

        if errors:
            summary.append("\n🔴 Errors:")
            summary.extend(errors)

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Barcode Upload Complete',
                    'message': '\n'.join(summary[:20]) + ('\n...(truncated)' if len(summary) > 20 else ''),
                    'type': 'success' if not errors else 'warning',
                    'sticky': False,
                },
            }
