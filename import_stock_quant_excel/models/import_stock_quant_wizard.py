from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
from io import BytesIO
import openpyxl

class ImportStockQuantWizard(models.TransientModel):
    _name = 'import.stock.quant.wizard'
    _description = 'Import Stock Quant Wizard'

    file = fields.Binary(required=True)
    filename = fields.Char()
    company_id = fields.Many2one('res.company', default=lambda self: self.env.company, required=True)
    state = fields.Selection([('draft','Draft'),('preview','Preview'),('done','Done')], default='draft')
    line_ids = fields.One2many('import.stock.quant.line','wizard_id')

    def action_preview(self):
        self.line_ids.unlink()
        data = base64.b64decode(self.file)
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        sheet = wb.active

        Product = self.env['product.product'].with_company(self.company_id)
        Location = self.env['stock.location']
        Lot = self.env['stock.lot']

        for row in range(2, sheet.max_row + 1):
            qty = float(sheet.cell(row,3).value or 0)
            vals = {
                'wizard_id': self.id,
                'row_no': row,
                'location_name': sheet.cell(row,1).value,
                'item_code': sheet.cell(row,2).value,
                'quantity': qty,
                'serial': sheet.cell(row,4).value,
                'is_valid': True,
            }

            if not vals['location_name'] or not vals['item_code'] or qty == 0:
                vals.update({'is_valid':False,'error_msg':'Missing data or zero qty'})
                self.env['import.stock.quant.line'].create(vals)
                continue

            location = Location.search([
                ('complete_name','=',vals['location_name']),
                ('usage','=','internal'),
                ('company_id','in',[self.company_id.id, False])
            ], limit=1)

            product = Product.search([('barcode','=',vals['item_code'])], limit=1)

            if not location or not product:
                vals.update({'is_valid':False,'error_msg':'Product or location not found'})
                self.env['import.stock.quant.line'].create(vals)
                continue

            vals.update({'location_id':location.id,'product_id':product.id})

            if vals['serial']:
                if product.tracking == 'serial' and abs(qty) != 1:
                    vals.update({'is_valid':False,'error_msg':'Serial qty must be 1 or -1'})
                lot = Lot.search([
                    ('name','=',vals['serial']),
                    ('product_id','=',product.id),
                    ('company_id','in',[self.company_id.id, False])
                ], limit=1)
                vals['lot_id'] = lot.id if lot else   vals.update({'is_valid':False,'error_msg':'Serial Not Found'})

            self.env['import.stock.quant.line'].create(vals)

        self.state = 'preview'

        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }
    def create_lots(self):
        for rec in self.line_ids:
            if rec.serial and not rec.lot_id:
                lot = self.env['stock.lot'].create({
                    'name': rec.serial,
                    'product_id': rec.product_id.id,
                    'company_id': rec.env.company.id,
                })
                rec.lot_id = lot.id
                rec.is_valid = True
                rec.error_msg = 'Serial Created'
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    def action_remove_invalid_lines(self):
        self.line_ids.filtered(lambda l: not l.is_valid).unlink()

    def action_apply(self):
        if self.line_ids.filtered(lambda l: not l.is_valid):
            raise UserError(_('Fix errors before applying'))

        Quant = self.env['stock.quant'].with_company(self.company_id)
        for line in self.line_ids:
            quant = Quant.search([
                ('product_id','=',line.product_id.id),
                ('location_id','=',line.location_id.id),
                ('lot_id','=',line.lot_id.id if line.lot_id else False),
            ], limit=1)

            if quant:
                if quant.quantity + line.quantity < 0:
                    raise UserError(_('Negative stock not allowed'))
                quant.quantity += line.quantity
            else:
                if line.quantity < 0:
                    raise UserError(_('No stock to subtract'))
                Quant.create({
                    'product_id': line.product_id.id,
                    'location_id': line.location_id.id,
                    'quantity': line.quantity,
                    'lot_id': line.lot_id.id if line.lot_id else False,
                    'company_id': self.company_id.id,
                })

        self.state = 'done'


class ImportStockQuantLine(models.TransientModel):
    _name = 'import.stock.quant.line'
    _description = 'Import Stock Quant Line'

    wizard_id = fields.Many2one('import.stock.quant.wizard', ondelete='cascade')
    row_no = fields.Integer()
    location_name = fields.Char()
    item_code = fields.Char()
    quantity = fields.Float()
    serial = fields.Char()
    product_id = fields.Many2one('product.product')
    location_id = fields.Many2one('stock.location')
    lot_id = fields.Many2one('stock.lot')
    is_valid = fields.Boolean(default=True)
    error_msg = fields.Text()

    def create_lot(self):
        if self.serial and not self.lot_id:
            lot = self.env['stock.lot'].create({
                'name': self.serial,
                'product_id': self.product_id.id,
                'company_id': self.env.company.id,
            })
            self.lot_id = lot.id
            self.is_valid = True
            self.error_msg = 'Serial Creadted'
        return {
            "type": "ir.actions.act_window",
            "res_model": self.wizard_id._name,
            "res_id": self.wizard_id.id,
            "view_mode": "form",
            "target": "new",
        }


