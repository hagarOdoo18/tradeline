# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import openpyxl
from io import BytesIO



class ProductPointWizard (models.TransientModel) :
    _name = 'product.point.wizard'

    date_from = fields.Date(
        string='From',
        required=False)

    date_to = fields.Date(
        string='TO',
        required=False)
    product_exel_Sheet = fields.Binary (string="Upload Product point Excel Sheet", )

    field = fields.Selection(
        string='Field',
        selection=[('point', 'Point'),
                   ('incentive', 'Incentive'), ],default='point',
        required=True, )


    def import_excel ( self ) :
        # Generating of the excel file to be read by openpyxl
        if self.product_exel_Sheet:
            try:
                data = base64.b64decode(self.product_exel_Sheet)
                wb = openpyxl.load_workbook(filename=BytesIO(data), data_only=True)
                sheet = wb.active
            except Exception as e:
                raise UserError(_('Invalid Excel file. Must be .xlsx format.\n\nError: %s') % str(e))

            headers = [str(cell.value).strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            header_map = {name: idx for idx, name in enumerate(headers) if name}

            # Create workbook

            for row in sheet.iter_rows(min_row=2):


                point = row[header_map['point']].value or ''
                item_code =str(row[header_map['item code']].value or '').strip()
                if item_code != None :
                    self.change_product_point(item_code,point)

        else:
            raise UserError (
                _ ("Upload Sheet"))

    def compute_product_point (self) :

        invoices_lines = self.sudo ().env ['account.move.line'].search (
            [('invoice_date', '>=', self.date_from),('invoice_date', '<=', self.date_to),('move_id.payment_state', 'in',['not_paid','paid','in_payment','partial','reversed'])])
        for line in invoices_lines :
            if line.move_id.move_type == "out_refund" :
                if self.field =='point':

                    line.product_point = line.product_id.product_point * -line.quantity

                else:
                    line.product_incentive = line.product_id.product_incentive * -line.quantity

            else :
                if self.field == 'point':

                    line.product_point = line.product_id.product_point * line.quantity

                else:
                    line.product_incentive = line.product_id.product_incentive * line.quantity

    def change_product_point ( self,item_code, point) :
        product_template = self.env['product.product'].search([('barcode', '=', item_code)])
        if not product_template:
            raise UserError(_("Please Check This Item [ "+item_code+" ]"))
        if self.field == 'point':
            product_template.product_point=point
        else:
            product_template.product_incentive=point





