# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import openpyxl
from io import BytesIO


class UploadDeliveryWizard(models.TransientModel):
    _name = 'upload.delivery.wizard'
    _description = 'Upload Excel to Delivery Lines'

    picking_id = fields.Many2one('stock.picking', string='Delivery', required=True)
    file = fields.Binary(string='Excel File', required=True)
    filename = fields.Char(string='File Name')
    create_if_not_exist = fields.Boolean(string='Create Serial if Not Exist', default=True)
    auto_confirm = fields.Boolean(string='Confirm Delivery Automatically', default=False)

    # ------------------------------------------------------------------
    # Main action
    # ------------------------------------------------------------------
    def action_upload_excel(self):
        from collections import defaultdict

        self.ensure_one()

        if not self.file:
            raise UserError(_("Please upload an Excel file."))

        # =============================
        # Load Excel
        # =============================
        try:
            data = base64.b64decode(self.file)
            wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
            sheet = wb.active
        except Exception as e:
            raise UserError(_("Invalid Excel file.\nError: %s") % str(e))

        # =============================
        # Headers
        # =============================
        headers = [str(c.value).strip().lower() for c in next(sheet.iter_rows(max_row=1))]
        header_map = {h: i for i, h in enumerate(headers) if h}

        for col in ('code', 'serial', 'quantity'):
            if col not in header_map:
                raise UserError(_("Missing column '%s' in Excel file.") % col)

        # =============================
        # Read rows
        # =============================
        rows = []
        barcodes = set()
        excel_serials = set()

        for r in sheet.iter_rows(min_row=2, values_only=True):
            code = str(r[header_map['code']] or '').strip()
            if not code:
                continue

            serial = str(r[header_map['serial']] or '').strip()
            qty = r[header_map['quantity']] or 0

            rows.append((code, serial, qty))
            barcodes.add(code)
            if serial:
                excel_serials.add(serial)

        if not rows:
            return {"type": "ir.actions.client", "tag": "reload"}

        picking = self.picking_id
        is_incoming = picking.picking_type_code == 'incoming'

        # PERFORMANCE: no chatter tracking / no auto-subscribe while we
        # rewrite thousands of operation lines.
        MoveLine = self.env['stock.move.line'].with_context(tracking_disable=True)

        # =============================
        # Batch fetch
        # =============================
        products = self.env['product.product'].search([('barcode', 'in', list(barcodes))])
        product_map = {p.barcode: p for p in products}

        # NOTE: the same product can sit on SEVERAL moves of one picking,
        # so keep every move per product instead of only the last one.
        moves_by_product = defaultdict(lambda: self.env['stock.move'])
        for mv in picking.move_ids_without_package:
            moves_by_product[mv.product_id.id] |= mv

        # prefetch every operation line of the picking in one query
        picking.move_line_ids_without_package.mapped('lot_id')

        lots = self.env['stock.lot'].sudo().search([
            ('name', 'in', list(excel_serials)),
            ('product_id', 'in', products.ids),
            ('company_id', 'in', [picking.company_id.id, False]),
        ])
        lot_map = {(l.product_id.id, l.name): l for l in lots}

        # =============================
        # Collect data (NO duplication)
        # =============================
        qty_map = defaultdict(float)   # product_id -> total qty
        serial_map = defaultdict(set)  # product_id -> set(serials)

        not_found_products = []
        errors = []

        for code, serial_name, quantity in rows:
            product = product_map.get(code)
            if not product:
                not_found_products.append(code)
                continue

            if product.tracking == 'serial':
                if not serial_name:
                    errors.append(_("Missing serial for product '%s'.") % code)
                    continue
                serial_map[product.id].add(serial_name)
            else:
                try:
                    qty_map[product.id] += float(quantity or 0)
                except Exception:
                    errors.append(_("Invalid quantity for product '%s'.") % code)

        processed = 0

        # PERFORMANCE: everything is accumulated and flushed in a handful of
        # SQL statements instead of one write()/create()/unlink() per serial.
        lines_to_unlink = MoveLine.browse()
        lines_to_pick = MoveLine.browse()
        move_line_vals = []

        # =============================================================
        # NON-serial products
        #
        # Odoo 17/18: stock.move.line no longer has `qty_done`; the done
        # quantity is `quantity` (+ `picked`). Writing `qty_done` on a
        # record was a silent no-op, so the reservation lines kept their
        # own quantity and the uploaded quantity was added on top of it
        # -> the doubled quantity on the move.
        # We OVERWRITE the existing line instead of adding a new one.
        # =============================================================
        for product_id, total_qty in qty_map.items():
            if total_qty <= 0:
                continue

            moves = moves_by_product.get(product_id)
            if not moves:
                product = self.env['product.product'].browse(product_id)
                errors.append(_("Product '%s' not found in picking.") % product.display_name)
                continue

            move = moves[0]
            lines = move.move_line_ids

            if lines:
                # keep exactly ONE line: the move quantity then equals the
                # file, never file + reservation.
                lines[0].write({'quantity': total_qty, 'picked': True})
                lines_to_unlink |= lines[1:]
            else:
                move_line_vals.append({
                    'move_id': move.id,
                    'picking_id': picking.id,
                    'product_id': product_id,
                    'product_uom_id': move.product_uom.id,
                    'quantity': total_qty,
                    'picked': True,
                    'location_id': move.location_id.id,
                    'location_dest_id': move.location_dest_id.id,
                })

            # other moves of the same product must not add quantity again
            for extra_move in moves[1:]:
                lines_to_unlink |= extra_move.move_line_ids

            processed += 1

        # =============================================================
        # SERIAL products
        # =============================================================

        # ---- Pass 1: create the missing lots (single DB call) ----
        missing_lot_vals = []
        missing_lot_keys = []

        if is_incoming and self.create_if_not_exist:
            for product_id, serials in serial_map.items():
                for serial_name in serials:
                    key = (product_id, serial_name)
                    if key in lot_map:
                        continue
                    missing_lot_keys.append(key)
                    missing_lot_vals.append({
                        'name': serial_name,
                        'product_id': product_id,
                        'company_id': picking.company_id.id,
                    })

        if missing_lot_vals:
            new_lots = self.env['stock.lot'].with_context(tracking_disable=True).create(missing_lot_vals)
            for key, lot in zip(missing_lot_keys, new_lots):
                lot_map[key] = lot

        # ---- Pass 2: exactly one move line per serial.
        # The empty reservation lines Odoo created when the picking was
        # reserved are dropped, otherwise their quantity is counted on top
        # of the uploaded serials -> the duplicated quantity. ----
        for product_id, serials in serial_map.items():
            moves = moves_by_product.get(product_id)
            if not moves:
                product = self.env['product.product'].browse(product_id)
                errors.append(_("Product '%s' not found in picking.") % product.display_name)
                continue

            move = moves[0]
            move_location_id = move.location_id.id
            move_dest_id = move.location_dest_id.id
            uom_id = move.product_uom.id

            already_there = set()
            for ml in moves.move_line_ids:
                if ml.lot_id:
                    already_there.add(ml.lot_id.name)
                    if ml.quantity != 1.0 or not ml.picked:
                        lines_to_pick |= ml
                else:
                    lines_to_unlink |= ml

            for serial_name in sorted(serials):
                if serial_name in already_there:
                    continue  # already on the picking -> never add it twice

                lot = lot_map.get((product_id, serial_name))
                if not lot:
                    product = self.env['product.product'].browse(product_id)
                    errors.append(
                        _("Serial '%s' not found for '%s'.") % (serial_name, product.display_name)
                    )
                    continue

                move_line_vals.append({
                    'move_id': move.id,
                    'picking_id': picking.id,
                    'product_id': product_id,
                    'product_uom_id': uom_id,
                    'lot_id': lot.id,
                    'quantity': 1.0,
                    'picked': True,
                    'location_id': move_location_id,
                    'location_dest_id': move_dest_id,
                })
                processed += 1

        # ---- Flush everything in three statements ----
        if lines_to_pick:
            lines_to_pick.write({'quantity': 1.0, 'picked': True})
        if lines_to_unlink:
            lines_to_unlink.unlink()
        if move_line_vals:
            MoveLine.create(move_line_vals)

        # =============================
        # Results
        # =============================
        messages = [_("Upload completed. Processed lines: %s") % processed]

        if not_found_products:
            messages.append(_("Products not found:"))
            messages.extend(sorted(set(not_found_products)))

        if errors:
            messages.append(_("Errors:"))
            messages.extend(errors)
            raise UserError("\n".join(messages))

        # =============================
        # Validation
        # =============================
        if self.auto_confirm:
            # button_validate may return the backorder confirmation popup;
            # return it instead of swallowing it, otherwise the picking
            # silently stays "Ready".
            action = self.action_auto_confirm()
            if isinstance(action, dict):
                return action

        return {
            "type": "ir.actions.client",
            "tag": "reload",
        }

    def action_auto_confirm(self):
        """Validate the picking now. Returns the backorder popup if Odoo asks
        for one, so the user keeps deciding what happens with the remainder."""
        self.ensure_one()
        try:
            return self.picking_id._upload_validate_now()
        except UserError:
            raise
        except Exception as e:
            raise UserError(_("Confirmation failed: %s") % str(e))
