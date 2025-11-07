from odoo import models, fields
from odoo.exceptions import UserError
import base64
from io import BytesIO
from openpyxl import load_workbook


class ProductAttributeImportWizard(models.TransientModel):
    _name = "product.attribute.import.wizard"
    _description = "Import Product Attributes from Excel"

    file = fields.Binary(string="Excel File", required=True)
    filename = fields.Char(string="Filename")

    def action_import_attributes(self):
        if not self.file:
            raise UserError("Please upload an Excel file.")

        try:
            file_data = base64.b64decode(self.file)
            workbook = load_workbook(BytesIO(file_data), data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise UserError(f"Invalid Excel file: {e}")

        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        required_cols = ["product_name", "attribute_name", "attribute_value"]

        if not all(col in headers for col in required_cols):
            raise UserError(
                "The Excel file must contain columns: product_name, attribute_name, attribute_value"
            )

        idx_product = headers.index("product_name")
        idx_attr = headers.index("attribute_name")
        idx_value = headers.index("attribute_value")

        ProductTemplate = self.env["product.template"]
        ProductAttribute = self.env["product.attribute"]
        ProductAttributeValue = self.env["product.attribute.value"]
        TemplateAttributeLine = self.env["product.template.attribute.line"]

        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            product_name = (row[idx_product] or "").strip() if row[idx_product] else ""
            attribute_name = (row[idx_attr] or "").strip() if row[idx_attr] else ""
            attribute_value_name = (row[idx_value] or "").strip() if row[idx_value] else ""

            if not (product_name and attribute_name and attribute_value_name):
                continue

            product_tmpl = ProductTemplate.search([("name", "=", product_name)], limit=1)
            if not product_tmpl:
                continue

            attribute = ProductAttribute.search([("name", "=", attribute_name)], limit=1)
            if not attribute:
                attribute = ProductAttribute.create({"name": attribute_name})

            attr_value = ProductAttributeValue.search([
                ("name", "=", attribute_value_name),
                ("attribute_id", "=", attribute.id)
            ], limit=1)
            if not attr_value:
                attr_value = ProductAttributeValue.create({
                    "name": attribute_value_name,
                    "attribute_id": attribute.id
                })

            line = TemplateAttributeLine.search([
                ("product_tmpl_id", "=", product_tmpl.id),
                ("attribute_id", "=", attribute.id)
            ], limit=1)

            if line:
                if attr_value.id not in line.value_ids.ids:
                    line.write({"value_ids": [(4, attr_value.id)]})
            else:
                TemplateAttributeLine.create({
                    "product_tmpl_id": product_tmpl.id,
                    "attribute_id": attribute.id,
                    "value_ids": [(6, 0, [attr_value.id])]
                })

            count += 1



