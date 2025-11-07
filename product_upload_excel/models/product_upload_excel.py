import base64
import openpyxl
from odoo import models, fields, _
from odoo.exceptions import UserError
from io import BytesIO
class ProductUploadExcel(models.TransientModel):
    _name = "product.upload.excel"
    _description = "Upload Products with Variants from Excel"

    file = fields.Binary(string="Excel File", required=True)
    filename = fields.Char(string="File Name")

    def action_import_products(self):
        if not self.file:
            raise UserError(_("Please upload an Excel file."))

        # Decode file
        data = base64.b64decode(self.file)
        workbook = openpyxl.load_workbook(filename=BytesIO(data), data_only=True)
        sheet = workbook.active

        # Expected headers
        header = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
        expected = ["Name", "Default Code", "Category", "Sales Price",  "Attribute", "Value",'Tracking','Company','Family']
        for h in expected:
            if h not in header:
                raise UserError(_("Invalid Excel format. Column %s not found.") % h)

        name_idx = header.index("Name")
        code_idx = header.index("Default Code")
        categ_idx = header.index("Category")
        family_idx = header.index("Family")
        price_idx = header.index("Sales Price")
        track_idx = header.index("Tracking")
        company_idx = header.index("Company")
        attr_idx = header.index("Attribute")
        val_idx = header.index("Value")

        products_map = {}

        # Read rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = str(row[name_idx] or "").strip()
            code = str(row[code_idx] or "").strip()
            categ_name = str(row[categ_idx] or "").strip()
            family_name = str(row[family_idx] or "").strip()
            company_name = str(row[company_idx] or "").strip()
            track_name = str(row[track_idx] or "").strip()
            price = float(row[price_idx] or 0.0)
            attr = str(row[attr_idx] or "").strip()
            val = str(row[val_idx] or "").strip()

            if not name or not attr or not val:
                continue

            if name not in products_map:
                products_map[name] = {
                    "code": code,
                    "category": categ_name,
                    "family": family_name,
                    "track_name": track_name,
                    "company_name": company_name,
                    "price": price,
                    "variants": {},
                }

            if attr not in products_map[name]["variants"]:
                products_map[name]["variants"][attr] = []
            products_map[name]["variants"][attr].append(val)
        company=False
        # Create products
        for pname, pdata in products_map.items():
            categ = self.env["product.category"].search([("name", "=", pdata["category"])], limit=1)
            if not categ:
                categ = self.env["product.category"].create({"name": pdata["category"]})

            company = self.env["res.company"].sudo().search([("name", "=", pdata["company_name"])], limit=1)
            family = self.env["product.family"].search([("name", "=", pdata["family"])], limit=1)
            if not family:
                 family = self.env["product.family"].create({"name": pdata["family"]})

            attribute_lines = []
            for attr, vals in pdata["variants"].items():
                attr_rec = self.env["product.attribute"].search([("name", "=", attr)], limit=1)
                if not attr_rec:
                    attr_rec = self.env["product.attribute"].create({"name": attr})

                value_ids = []
                for v in set(vals):
                    val_rec = self.env["product.attribute.value"].search([
                        ("name", "=", v), ("attribute_id", "=", attr_rec.id)
                    ], limit=1)
                    if not val_rec:
                        val_rec = self.env["product.attribute.value"].create({
                            "name": v,
                            "attribute_id": attr_rec.id
                        })
                    value_ids.append(val_rec.id)

                attribute_lines.append((0, 0, {"attribute_id": attr_rec.id, "value_ids": [(6, 0, value_ids)]}))

            template= self.env["product.template"].create({
                "name": pname,
                "default_code": pdata["code"],
                "categ_id": categ.id,
                "company_id": company.id if company else False,
                "family_id": family.id,
                "available_in_pos": True,
                "is_storable": True,
                "invoice_policy": 'delivery',
                "tracking": pdata['track_name'],
                "list_price": pdata["price"],
                "attribute_line_ids": attribute_lines,
            })
            # barcode_idx = header.index("Barcode")
            #
            # template._create_variant_ids()
            #
            # # عدّي على كل Variant وحط الكود
            #
            # for variant in template.product_variant_ids:
            #     # نحاول نلاقي الصف اللي يخص الـ variant الحالي
            #     attrs = variant.product_template_attribute_value_ids.mapped("name")
            #     for row in sheet.iter_rows(min_row=2, values_only=True):
            #         # لو كل الـ attribute values متطابقة
            #         variant_code = str(row[barcode_idx] or "").strip()
            #         val = str(row[val_idx] or "").strip()
            #         if val in attrs:
            #             if not variant.barcode:
            #                 variant.barcode = variant_code
